"""
SSTQA benchmark runner.

Dataset: OpenDataBox/ST-Raptor SSTQA-en (GitHub)
  - 102 semi-structured tables as Excel files
  - Questions in JSONL format (id, table_id, query, label)
Primary metrics: Accuracy, ROUGE-L, ANLS, F1

Architecture:
  - Each Excel table is parsed to text (preserving structure)
  - Table text is chunked with RecursiveCharacterTextSplitter
  - When use_table_parent_child=true: rows become child chunks, full table
    stored as parent. At retrieval, child hits are swapped for parent text.
  - When build_raptor_tree=true: RAPTOR tree is built per table
  - For each question: embed query -> cosine search -> top-K -> LLM answer
  - Scoring: max over ground truths (single GT per question in SSTQA)

Downloads tables and questions from the ST-Raptor GitHub repository.
"""

import json
import logging
import os
import ssl
import urllib.request
from pathlib import Path
from uuid import uuid4

import numpy as np
from langchain_text_splitters import RecursiveCharacterTextSplitter

from benchmarks.base_benchmark import BaseBenchmark
from metrics import (
    anls_score,
    exact_match,
    f1_token_score,
    rouge_l_score,
    score_with_multiple_gts,
)

logger = logging.getLogger(__name__)

# GitHub raw URLs for SSTQA-en dataset
_GITHUB_BASE = "https://raw.githubusercontent.com/OpenDataBox/ST-Raptor/master/data/SSTQA-en"
_TEST_JSONL_URL = f"{_GITHUB_BASE}/test.jsonl"

CHUNK_SIZE = 1600
CHUNK_OVERLAP = 200
DEFAULT_TOP_K = 5

_EMBED_CACHE_DIR = "sstqa_embeddings"
_TREE_CACHE_DIR = "sstqa_trees"

QA_GENERATION_PROMPT = """Answer the question based on the table context below.
Give a concise, direct answer. If the answer is a number, include the unit if present.
If the information is not in the context, say "unanswerable".

Context:
{context}

Question: {question}

Answer:"""


def _download_file(url: str, dest: Path) -> bool:
    """Download a file if not already cached. Returns True on success."""
    if dest.exists():
        return True
    dest.parent.mkdir(parents=True, exist_ok=True)
    logger.info(f"Downloading {url}...")
    try:
        import certifi
        ssl_ctx = ssl.create_default_context(cafile=certifi.where())
    except ImportError:
        ssl_ctx = ssl.create_default_context()
    req = urllib.request.Request(url)
    try:
        with urllib.request.urlopen(req, context=ssl_ctx) as response:
            with open(dest, "wb") as f:
                while True:
                    chunk = response.read(8192)
                    if not chunk:
                        break
                    f.write(chunk)
        return True
    except urllib.error.HTTPError as e:
        logger.warning(f"Download failed ({e.code}): {url}")
        return False


def _list_github_table_files() -> list[int]:
    """Get list of available table IDs from GitHub API."""
    api_url = "https://api.github.com/repos/OpenDataBox/ST-Raptor/contents/data/SSTQA-en/table"
    try:
        import certifi
        ssl_ctx = ssl.create_default_context(cafile=certifi.where())
    except ImportError:
        ssl_ctx = ssl.create_default_context()
    req = urllib.request.Request(api_url, headers={"Accept": "application/json"})
    with urllib.request.urlopen(req, context=ssl_ctx) as response:
        items = json.loads(response.read())
    table_ids = []
    for item in items:
        name = item.get("name", "")
        if name.endswith(".xlsx"):
            try:
                table_ids.append(int(name.replace(".xlsx", "")))
            except ValueError:
                pass
    return sorted(table_ids)


def _parse_excel_to_text(xlsx_path: Path) -> str:
    """Parse an Excel file into a structured text representation.

    Preserves table structure by rendering each sheet as a pipe-delimited
    table with headers, handling merged cells where possible.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "Install openpyxl: pip install openpyxl"
        )

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sections: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(cells):
                rows.append(cells)

        if not rows:
            continue

        # Build pipe-delimited table representation
        lines: list[str] = []
        if len(wb.sheetnames) > 1:
            lines.append(f"Sheet: {sheet_name}")

        for row_cells in rows:
            lines.append("| " + " | ".join(row_cells) + " |")

        sections.append("\n".join(lines))

    wb.close()
    return "\n\n".join(sections)


def _cosine_similarity_batch(query: np.ndarray, matrix: np.ndarray) -> np.ndarray:
    """Cosine similarity between a query vector and a matrix of vectors."""
    q_norm = np.linalg.norm(query)
    if q_norm < 1e-10:
        return np.zeros(matrix.shape[0])
    q_hat = query / q_norm
    m_norms = np.linalg.norm(matrix, axis=1, keepdims=True)
    m_norms = np.maximum(m_norms, 1e-10)
    m_hat = matrix / m_norms
    return m_hat @ q_hat


class SstqaBenchmark(BaseBenchmark):

    def __init__(self, config_path: str):
        super().__init__(config_path)
        self._build_raptor = self.config.get("build_raptor_tree", False)
        self._use_table_pc = self.config.get("use_table_parent_child", False)
        self._use_expansion = self.config.get("use_retrieval_expansion", False)
        self._top_k = self.config.get("top_k", DEFAULT_TOP_K)
        self._splitter = RecursiveCharacterTextSplitter(
            chunk_size=CHUNK_SIZE, chunk_overlap=CHUNK_OVERLAP
        )
        # Per-table indices: table_id -> {chunks, embeddings, parent_ids, ...}
        self._table_indices: dict[int, dict] = {}
        self._parent_stores: dict[int, dict[str, str]] = {}
        self._data_dir: str = "datasets"

    def load_dataset(self, data_dir: str = "datasets") -> None:
        self._data_dir = data_dir
        cache_dir = Path(data_dir) / "sstqa_raw"

        # Download test questions
        test_path = cache_dir / "test.jsonl"
        _download_file(_TEST_JSONL_URL, test_path)

        # Load questions
        questions = []
        with open(test_path) as f:
            for line in f:
                line = line.strip()
                if line:
                    questions.append(json.loads(line))

        logger.info(f"Loaded {len(questions)} SSTQA questions")

        # Apply sample_size and sharding BEFORE downloading tables
        sample_size = self.config.get("sample_size")
        shard_id = self.config.get("shard_id")
        num_shards = self.config.get("num_shards")

        if sample_size and sample_size < len(questions):
            questions = questions[:sample_size]
            logger.info(f"Limited to {sample_size} questions")

        if shard_id is not None and num_shards is not None:
            questions = [
                q for i, q in enumerate(questions)
                if i % num_shards == shard_id
            ]
            logger.info(f"Shard {shard_id}/{num_shards}: {len(questions)} questions")

        # Download only needed table Excel files
        needed_table_ids = {q["table_id"] for q in questions}
        logger.info(f"Need {len(needed_table_ids)} tables")

        table_dir = cache_dir / "table"
        table_dir.mkdir(parents=True, exist_ok=True)

        available_tables: set[int] = set()
        for tid in sorted(needed_table_ids):
            xlsx_path = table_dir / f"{tid}.xlsx"
            xlsx_url = f"{_GITHUB_BASE}/table/{tid}.xlsx"
            if _download_file(xlsx_url, xlsx_path):
                available_tables.add(tid)

        # Filter out questions whose tables couldn't be downloaded
        if len(available_tables) < len(needed_table_ids):
            missing = needed_table_ids - available_tables
            logger.warning(f"Skipping questions for unavailable tables: {sorted(missing)}")
            questions = [q for q in questions if q["table_id"] in available_tables]

        # Build dataset
        self.dataset = []
        for q in questions:
            self.dataset.append({
                "question": q["query"],
                "answers": [q["label"]],
                "kwargs": {"table_id": q["table_id"]},
            })

        # Pre-build table indices (chunk + embed all needed tables)
        final_table_ids = {item["kwargs"]["table_id"] for item in self.dataset}
        for tid in sorted(final_table_ids):
            self._ensure_table_index(tid, table_dir)

        logger.info(
            f"SSTQA ready: {len(self.dataset)} questions, "
            f"{len(self._table_indices)} tables indexed"
        )

    def _ensure_table_index(self, table_id: int, table_dir: Path) -> None:
        """Build chunk index for a table if not already cached."""
        if table_id in self._table_indices:
            return

        from modules.embeddings import embed, embed_batch

        embed_cache_dir = Path(self._data_dir) / _EMBED_CACHE_DIR
        embed_cache_dir.mkdir(parents=True, exist_ok=True)
        cache_file = embed_cache_dir / f"table_{table_id}.npz"

        xlsx_path = table_dir / f"{table_id}.xlsx"
        table_text = _parse_excel_to_text(xlsx_path)

        if not table_text.strip():
            logger.warning(f"Table {table_id}: empty after parsing")
            self._table_indices[table_id] = {
                "chunks": [], "embeddings": np.zeros((0, 3072))
            }
            return

        # Generate chunks
        chunks: list[str] = []
        parent_ids: list[str] = []

        if self._use_table_pc:
            # Table P-C: each row line is a child chunk, full table is parent
            parent_id = str(uuid4())
            self._parent_stores.setdefault(table_id, {})[parent_id] = table_text

            lines = table_text.split("\n")
            for line in lines:
                line = line.strip()
                if line and line.startswith("|"):
                    chunks.append(line)
                    parent_ids.append(parent_id)

            # Also add the full table as a chunk for flat comparison
            if not chunks:
                chunks = self._splitter.split_text(table_text)
                parent_ids = [""] * len(chunks)
        else:
            chunks = self._splitter.split_text(table_text)
            parent_ids = [""] * len(chunks)

        if not chunks:
            self._table_indices[table_id] = {
                "chunks": [], "embeddings": np.zeros((0, 3072))
            }
            return

        # Load or compute embeddings
        if cache_file.exists():
            data = np.load(cache_file, allow_pickle=True)
            cached_chunks = list(data["chunks"])
            cached_embeddings = data["embeddings"]
            if cached_chunks == chunks:
                self._table_indices[table_id] = {
                    "chunks": chunks,
                    "embeddings": cached_embeddings,
                    "parent_ids": parent_ids,
                }
                return

        # Embed all chunks
        logger.info(f"Embedding {len(chunks)} chunks for table {table_id}")
        raw_embeddings = embed_batch(chunks)
        dim = 3072
        emb_matrix = np.zeros((len(chunks), dim), dtype=np.float32)
        for i, emb in enumerate(raw_embeddings):
            if emb is not None:
                emb_matrix[i] = np.array(emb, dtype=np.float32)

        # Cache
        np.savez(cache_file, chunks=np.array(chunks, dtype=object), embeddings=emb_matrix)

        # Build RAPTOR tree if configured
        if self._build_raptor:
            tree_cache_dir = Path(self._data_dir) / _TREE_CACHE_DIR
            tree_cache_dir.mkdir(parents=True, exist_ok=True)
            tree_file = tree_cache_dir / f"table_{table_id}.npz"

            if tree_file.exists():
                tree_data = np.load(tree_file, allow_pickle=True)
                summary_chunks = list(tree_data["chunks"])
                summary_embeddings = tree_data["embeddings"]
            else:
                from benchmarks.raptor_tree import build_raptor_tree
                summary_chunks, summary_embeddings = build_raptor_tree(
                    chunks, emb_matrix
                )
                if summary_chunks:
                    np.savez(
                        tree_file,
                        chunks=np.array(summary_chunks, dtype=object),
                        embeddings=summary_embeddings,
                    )

            if summary_chunks:
                chunks = chunks + summary_chunks
                parent_ids = parent_ids + [""] * len(summary_chunks)
                emb_matrix = np.vstack([emb_matrix, summary_embeddings])

        self._table_indices[table_id] = {
            "chunks": chunks,
            "embeddings": emb_matrix,
            "parent_ids": parent_ids,
        }

    def _retrieve(self, question: str, table_id: int) -> list[str]:
        """Retrieve top-K chunks for a question from the table index."""
        from modules.embeddings import embed

        index = self._table_indices.get(table_id)
        if index is None or len(index["chunks"]) == 0:
            return []

        query_emb = embed(text=question)
        if query_emb is None:
            logger.error(f"Failed to embed question: {question[:80]}")
            return []
        query_vec = np.array(query_emb, dtype=np.float32)
        scores = _cosine_similarity_batch(query_vec, index["embeddings"])
        top_indices = np.argsort(scores)[::-1][:self._top_k]

        if self._use_expansion and self._use_table_pc:
            chunk_pids = index.get("parent_ids", [""] * len(index["chunks"]))
            retrieved = [
                {"text": index["chunks"][k], "parent_id": chunk_pids[k]}
                for k in top_indices
            ]
            return self._expand_parents(retrieved, table_id)

        return [index["chunks"][k] for k in top_indices]

    def _expand_parents(
        self, retrieved: list[dict], table_id: int
    ) -> list[str]:
        """Swap table row child hits for their full parent table text."""
        parent_store = self._parent_stores.get(table_id, {})
        if not parent_store:
            return [n["text"] for n in retrieved]

        expanded: list[str] = []
        seen_parents: set[str] = set()

        for node in retrieved:
            pid = node.get("parent_id", "")
            if pid and pid in parent_store:
                if pid not in seen_parents:
                    expanded.append(parent_store[pid])
                    seen_parents.add(pid)
            else:
                expanded.append(node["text"])

        return expanded

    def run_single(self, question: str, table_id: int = 0, **kwargs) -> str:
        """Retrieve context chunks, generate answer with Gemini Flash."""
        from core.config import generate_with_retry, MODEL_FLASH
        from google.genai import types as genai_types

        chunks = self._retrieve(question, table_id)
        if not chunks:
            return "unanswerable"

        context = "\n\n---\n\n".join(chunks)
        prompt = QA_GENERATION_PROMPT.format(context=context, question=question)

        try:
            response = generate_with_retry(
                model=MODEL_FLASH,
                contents=prompt,
                config=genai_types.GenerateContentConfig(
                    temperature=0.0,
                    max_output_tokens=256,
                ),
            )
            return response.text.strip() if response.text else "unanswerable"
        except Exception as e:
            logger.error(f"Generation failed for question '{question[:60]}': {e}")
            return "unanswerable"

    def evaluate(
        self,
        predictions: list[str],
        ground_truths: list[list[str]],
    ) -> dict[str, float]:
        acc_scores = [
            score_with_multiple_gts(pred, gts, exact_match)
            for pred, gts in zip(predictions, ground_truths)
        ]
        rouge_scores = [
            score_with_multiple_gts(pred, gts, rouge_l_score)
            for pred, gts in zip(predictions, ground_truths)
        ]
        anls_scores = [
            score_with_multiple_gts(pred, gts, anls_score)
            for pred, gts in zip(predictions, ground_truths)
        ]
        f1_scores = [
            score_with_multiple_gts(pred, gts, f1_token_score)
            for pred, gts in zip(predictions, ground_truths)
        ]
        n = max(len(acc_scores), 1)
        return {
            "accuracy": sum(acc_scores) / n,
            "rouge_l": sum(rouge_scores) / n,
            "anls": sum(anls_scores) / n,
            "f1": sum(f1_scores) / n,
        }
